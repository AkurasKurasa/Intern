"""Build the synthetic grade sheets for the mock portal experiment.

Source template: the AUF "E-Grade-Book (new trans.)" workbook, an institutional
grade book with a MASTERLIST, per-term score sheets, a SUMMARY of transmuted
grades, and a Reference transmutation table. The layout, formatting, merged
header block and transmutation table are kept exactly; every identity and every
score is replaced with synthetic data.

Two deliberate changes to the template, both required by the experiment:

  1. PROGRAM and YEAR LEVEL columns are added. The real book carries course and
     term in the header block, not per student, which would leave the portal's
     Course and Year fields with no source column at all.
  2. Formulas are replaced with their computed values. openpyxl cannot write a
     cached result alongside a formula, so a formula-bearing sheet reads back as
     empty under pandas.read_excel - which is how the pipeline consumes it.

Student identity comes from mocksite/shared/roster.js, whose Student IDs are the
join key the portal matches on. Nothing here may invent an ID.

Usage:  python data/sheets/make_sheets.py
"""

import random
import re
import shutil
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

REPO = Path(__file__).resolve().parents[2]
ROSTER_JS = REPO.parents[1] / "practice_apps" / "mocksite" / "shared" / "roster.js"
OUT_DIR = REPO / "data" / "sheets"

TEMPLATE_LEC = Path.home() / "Downloads" / "21-09-14-E-Grade-Book-lec-new-trans. (1).xlsx"

SEED = 20260805

# Header block, MASTERLIST column B. The real workbook names a real institution
# and a real instructor; both are replaced.
HEADER_BLOCK = {
    2: "MOCK STATE UNIVERSITY",
    3: "Sample City",
    5: "COLLEGE OF COMPUTER STUDIES",
    6: "DEPARTMENT OF COMPUTER SCIENCE",
    7: "1st Semester, AY 2021 - 2022",
    9: "MASTERLIST",
    10: "CS301: DATA STRUCTURES AND ALGORITHMS",
    11: "MWF / 10:00-11:30am / Online",
    12: "PROF. MARIA S. DELA CRUZ",
}

PROGRAMS = [
    "BS Computer Science",
    "BS Information Technology",
    "BS Information Systems",
]

# MASTERLIST: data rows, and the two columns added for the experiment.
ML_FIRST_ROW = 17
ML_COL_NO, ML_COL_ID = 2, 3
ML_COL_LAST, ML_COL_FIRST, ML_COL_MI = 4, 5, 6
ML_COL_PROGRAM, ML_COL_YEAR = 7, 8

# Term sheets (MIDTERM / FINALS): header at row 12-14, maxima at row 16, data
# from row 17. F:O in-process assessments, P:Y terminal assessments, AA major
# exam, Z class standing, AB exam equivalent, AC raw grade, AD transmuted
# grade, AE remarks.
TERM_FIRST_ROW = 17
TERM_MAX_ROW = 16
COL_IN_FIRST, COL_IN_LAST = 6, 15      # F:O
COL_TERM_FIRST, COL_TERM_LAST = 16, 25  # P:Y
COL_CS, COL_EXAM, COL_EQUI = 26, 27, 28  # Z, AA, AB
COL_RAW, COL_GRADE, COL_REMARKS = 29, 30, 31  # AC, AD, AE

ACTIVITY_MAX = 10
EXAM_MAX = 100

# SUMMARY: header row 12, data from row 15.
SUM_FIRST_ROW = 15
SUM_COL_NO, SUM_COL_ID = 1, 2
SUM_COL_LAST, SUM_COL_FIRST, SUM_COL_MI = 3, 4, 5
SUM_COL_MIDTERM, SUM_COL_FINAL = 6, 7
SUM_COL_PROGRAM, SUM_COL_YEAR, SUM_COL_FINAL_GRADE = 8, 9, 10


def read_roster():
    """Student IDs and names from roster.js, in roster order."""
    text = ROSTER_JS.read_text(encoding="utf-8")
    pairs = re.findall(
        r'student_id:\s*"([^"]+)"\s*,\s*student_name:\s*"([^"]+)"', text
    )
    if not pairs:
        raise SystemExit(f"no roster entries parsed from {ROSTER_JS}")

    students = []
    for sid, name in pairs:
        # "Abad, Andrea A." -> last, first, middle initial
        last, rest = name.split(",", 1)
        rest = rest.strip().split()
        mi = rest[-1] if rest and rest[-1].endswith(".") else ""
        first = " ".join(rest[:-1] if mi else rest)
        students.append(
            {
                "id": sid,
                "last": last.strip().upper(),
                "first": first.upper(),
                "mi": mi.upper(),
                "display": name,
            }
        )
    return students


def read_transmutation(wb):
    """The Reference table as (raw_cutoff, transmuted) pairs, ascending."""
    ws = wb["Reference"]
    table = []
    for r in range(4, 40):
        raw, grade = ws.cell(r, 2).value, ws.cell(r, 3).value
        if isinstance(raw, (int, float)) and isinstance(grade, (int, float)):
            table.append((float(raw), float(grade)))
    table.sort()
    if not table:
        raise SystemExit("Reference transmutation table not found")
    return table


def transmute(raw, table):
    """Excel LOOKUP against the transmutation table: last cutoff <= raw."""
    out = table[0][1]
    for cutoff, grade in table:
        if raw >= cutoff:
            out = grade
        else:
            break
    return out


def put(ws, row, col, value):
    """Write unless the target is a merged continuation cell, which openpyxl
    makes read-only. The template's footer rows carry merged summary bands."""
    cell = ws.cell(row, col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        return False
    cell.value = value
    return True


# On MIDTERM / FINALS / SUMMARY the title band is only these rows. Row 9 is the
# sheet's own title and rows 12-14 are the column headers - writing the
# MASTERLIST block over those would destroy 'No.', 'STUDENT NUMBER' and the rest.
SHARED_HEADER_ROWS = (2, 3, 5, 6, 7, 10)


def rewrite_header_block(ws, col=1):
    """Replace the institution/course title band on a non-MASTERLIST sheet.

    Some of these cells are '=MASTERLIST!Bn' references and some are literal
    text repeated per sheet; both have to go, or the source institution leaks
    into the artifact.
    """
    for row in SHARED_HEADER_ROWS:
        cell = ws.cell(row, col)
        if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = HEADER_BLOCK[row]


def clear_below(ws, first_row, cols):
    for r in range(first_row, ws.max_row + 1):
        for c in cols:
            put(ws, r, c, None)


def unprotect(wb):
    """Drop the template's sheet protection.

    The source workbook ships locked, and copying it carries the lock over.
    That is right for a live grade book and wrong for a test fixture: a
    demonstrator has to be able to click and copy cells freely, and Excel
    refuses with "the cell you're trying to change is on a protected sheet".
    """
    for ws in wb.worksheets:
        ws.protection.sheet = False
        ws.protection.enable()
        ws.protection.sheet = False
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = False


def build(template, out_path, rng):
    shutil.copy(template, out_path)
    wb = openpyxl.load_workbook(out_path)
    unprotect(wb)
    table = read_transmutation(wb)
    students = read_roster()

    # ---- MASTERLIST -----------------------------------------------------
    ml = wb["MASTERLIST"]
    for row, text in HEADER_BLOCK.items():
        ml.cell(row, 2).value = text

    ml.cell(14, ML_COL_PROGRAM).value = "PROGRAM"
    ml.cell(14, ML_COL_YEAR).value = "YEAR LEVEL"

    clear_below(ml, ML_FIRST_ROW, range(2, 9))
    for i, s in enumerate(students):
        r = ML_FIRST_ROW + i
        s["program"] = rng.choice(PROGRAMS)
        s["year"] = rng.choice([1, 2, 3, 3, 4])
        ml.cell(r, ML_COL_NO).value = i + 1
        ml.cell(r, ML_COL_ID).value = s["id"]
        ml.cell(r, ML_COL_LAST).value = s["last"]
        ml.cell(r, ML_COL_FIRST).value = s["first"]
        ml.cell(r, ML_COL_MI).value = s["mi"]
        ml.cell(r, ML_COL_PROGRAM).value = s["program"]
        ml.cell(r, ML_COL_YEAR).value = s["year"]

    # ---- term sheets ----------------------------------------------------
    # Ability per student is fixed across terms, so midterm and final grades
    # correlate the way real ones do instead of being independent noise.
    for s in students:
        s["ability"] = rng.gauss(0.78, 0.13)
        s["grades"] = {}

    for term in ("MIDTERM", "FINALS"):
        ws = wb[term]
        rewrite_header_block(ws)

        for c in range(COL_IN_FIRST, COL_TERM_LAST + 1):
            ws.cell(TERM_MAX_ROW, c).value = ACTIVITY_MAX
        ws.cell(TERM_MAX_ROW, COL_EXAM).value = EXAM_MAX

        clear_below(ws, TERM_FIRST_ROW, range(1, COL_REMARKS + 1))

        for i, s in enumerate(students):
            r = TERM_FIRST_ROW + i
            ws.cell(r, 1).value = i + 1
            ws.cell(r, 2).value = s["id"]
            ws.cell(r, 3).value = s["last"]
            ws.cell(r, 4).value = s["first"]
            ws.cell(r, 5).value = s["mi"]

            def score():
                v = rng.gauss(s["ability"], 0.10) * ACTIVITY_MAX
                return round(min(ACTIVITY_MAX, max(0.0, v)))

            in_scores = [score() for _ in range(COL_IN_LAST - COL_IN_FIRST + 1)]
            term_scores = [score() for _ in range(COL_TERM_LAST - COL_TERM_FIRST + 1)]
            for n, v in enumerate(in_scores):
                ws.cell(r, COL_IN_FIRST + n).value = v
            for n, v in enumerate(term_scores):
                ws.cell(r, COL_TERM_FIRST + n).value = v

            exam = round(min(EXAM_MAX, max(0.0, rng.gauss(s["ability"], 0.11) * EXAM_MAX)))
            ws.cell(r, COL_EXAM).value = exam

            in_pct = sum(in_scores) / (ACTIVITY_MAX * len(in_scores))
            term_pct = sum(term_scores) / (ACTIVITY_MAX * len(term_scores))
            cs = round((in_pct * 0.4 + term_pct * 0.6) * 60, 3)
            equi = round((exam / EXAM_MAX) * 40, 3)
            raw = round(cs + equi, 3)
            grade = transmute(raw, table)

            ws.cell(r, COL_CS).value = cs
            ws.cell(r, COL_EQUI).value = equi
            ws.cell(r, COL_RAW).value = raw
            ws.cell(r, COL_GRADE).value = grade
            ws.cell(r, COL_REMARKS).value = "PASSED" if grade >= 75 else "FAILED"

            s["grades"][term] = grade

    # ---- SUMMARY --------------------------------------------------------
    sm = wb["SUMMARY"]
    rewrite_header_block(sm)
    sm.cell(9, 1).value = "CLASS RECORD (SUMMARY OF GRADES)"

    sm.cell(12, SUM_COL_PROGRAM).value = "PROGRAM"
    sm.cell(12, SUM_COL_YEAR).value = "YEAR LEVEL"
    sm.cell(12, SUM_COL_FINAL_GRADE).value = "FINAL GRADE"

    clear_below(sm, SUM_FIRST_ROW, range(1, SUM_COL_FINAL_GRADE + 1))
    for i, s in enumerate(students):
        r = SUM_FIRST_ROW + i
        final_grade = round((s["grades"]["MIDTERM"] + s["grades"]["FINALS"]) / 2)
        sm.cell(r, SUM_COL_NO).value = i + 1
        sm.cell(r, SUM_COL_ID).value = s["id"]
        sm.cell(r, SUM_COL_LAST).value = s["last"]
        sm.cell(r, SUM_COL_FIRST).value = s["first"]
        sm.cell(r, SUM_COL_MI).value = s["mi"]
        sm.cell(r, SUM_COL_MIDTERM).value = s["grades"]["MIDTERM"]
        sm.cell(r, SUM_COL_FINAL).value = s["grades"]["FINALS"]
        sm.cell(r, SUM_COL_PROGRAM).value = s["program"]
        sm.cell(r, SUM_COL_YEAR).value = s["year"]
        sm.cell(r, SUM_COL_FINAL_GRADE).value = final_grade
        s["final_grade"] = final_grade

    wb.save(out_path)
    return students


def add_status_column(src, dst):
    """The 3.11 'Status column present' condition: same sheet, plus a
    precomputed REMARKS column on SUMMARY. Remarks should then be *mapped*
    rather than derived, and feature 17 is what should fire."""
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    unprotect(wb)
    sm = wb["SUMMARY"]
    col = SUM_COL_FINAL_GRADE + 1
    sm.cell(12, col).value = "REMARKS"
    for r in range(SUM_FIRST_ROW, SUM_FIRST_ROW + 50):
        grade = sm.cell(r, SUM_COL_FINAL_GRADE).value
        if grade is None:
            break
        sm.cell(r, col).value = "PASSED" if grade >= 75 else "FAILED"
    wb.save(dst)


def main():
    if not TEMPLATE_LEC.exists():
        raise SystemExit(f"template not found: {TEMPLATE_LEC}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    rng = random.Random(SEED)
    base = OUT_DIR / "grade_sheet.xlsx"
    students = build(TEMPLATE_LEC, base, rng)
    add_status_column(base, OUT_DIR / "grade_sheet_status.xlsx")

    passed = sum(1 for s in students if s["final_grade"] >= 75)
    lo = min(s["final_grade"] for s in students)
    hi = max(s["final_grade"] for s in students)
    print(f"wrote {base.name} and grade_sheet_status.xlsx")
    print(f"  {len(students)} students, final grade {lo}-{hi}, "
          f"{passed} passed / {len(students) - passed} failed")


if __name__ == "__main__":
    main()
