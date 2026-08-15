"""Check the generated sheets against what the pipeline assumes.

Fails loudly rather than reporting a green run on a broken artifact.
Usage:  python data/sheets/verify_sheets.py
"""

import re
import sys
import warnings
from pathlib import Path

import pandas as pd

warnings.filterwarnings("ignore")

REPO = Path(__file__).resolve().parents[2]
SHEETS = REPO / "data" / "sheets"
ROSTER_JS = REPO.parents[1] / "practice_apps" / "mocksite" / "shared" / "roster.js"

# SUMMARY's header block occupies rows 1-14; the column names sit on row 12 and
# the data starts on row 15. 3.1's "row 1 is the header" default does not hold
# here - this is the one-time user override the spec calls for.
SUMMARY_HEADER_ROW = 12

failures = []


def check(label, condition, detail=""):
    mark = "ok  " if condition else "FAIL"
    print(f"  [{mark}] {label}" + (f" - {detail}" if detail else ""))
    if not condition:
        failures.append(label)


def roster_ids():
    text = ROSTER_JS.read_text(encoding="utf-8")
    return re.findall(r'student_id:\s*"([^"]+)"', text)


def load_summary(path):
    df = pd.read_excel(path, sheet_name="SUMMARY", header=SUMMARY_HEADER_ROW - 1)
    df = df.dropna(how="all")
    df = df[df.iloc[:, 1].notna()]
    return df


def main():
    ids = roster_ids()
    base = SHEETS / "grade_sheet.xlsx"
    status = SHEETS / "grade_sheet_status.xlsx"

    for path in (base, status):
        if not path.exists():
            sys.exit(f"missing {path} - run make_sheets.py first")

    print(f"\n{base.name}")
    df = load_summary(base)
    cols = [str(c) for c in df.columns]
    check("50 data rows", len(df) == 50, f"got {len(df)}")
    check("student IDs match the roster exactly",
          list(df.iloc[:, 1].astype(str)) == ids)

    for name in ("No.", "STUDENT NUMBER", "NAME OF STUDENT",
                 "PROGRAM", "YEAR LEVEL", "FINAL GRADE", "MIDTERM", "FINAL"):
        check(f"column {name!r} present", any(name == c.strip() for c in cols),
              f"columns={cols}")

    # The title band and the column-header row are adjacent; writing one over
    # the other is a silent corruption, so pin the template's own headers.
    check("template column headers survived the header-block rewrite",
          cols[:3] == ["No.", "STUDENT NUMBER", "NAME OF STUDENT"],
          f"got {cols[:3]}")

    grades = pd.to_numeric(df["FINAL GRADE"], errors="coerce")
    check("final grades all numeric", grades.notna().all())
    check("grades within the transmuted range 65-100",
          bool(grades.min() >= 65 and grades.max() <= 100),
          f"{grades.min()}-{grades.max()}")
    n_pass = int((grades >= 75).sum())
    n_fail = int((grades < 75).sum())
    check("both classes present either side of 75 (3.8 needs a failing row)",
          n_pass > 0 and n_fail > 0, f"{n_pass} passed / {n_fail} failed")

    check("no REMARKS column - this is the 'derived' condition",
          not any("REMARK" in c.upper() for c in cols))

    print("\nno formulas left anywhere (pandas reads values, not '=')")
    import openpyxl
    wb = openpyxl.load_workbook(base)
    stray = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    stray.append(f"{ws.title}!{cell.coordinate}")
    check("zero formula cells", not stray, f"{len(stray)} found: {stray[:5]}")

    print("\nreal identities removed")
    text = " ".join(
        str(c.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if c.value is not None
    ).upper()
    for token in ("ANGELES", "TAGUINES", "JERRENCE", "06-0920-662"):
        check(f"{token!r} absent", token not in text)

    print(f"\n{status.name}")
    dfs = load_summary(status)
    cols_s = [str(c).strip() for c in dfs.columns]
    check("REMARKS column present - this is the 'mapped' condition",
          "REMARKS" in cols_s, f"columns={cols_s}")
    if "REMARKS" in cols_s:
        vals = set(dfs["REMARKS"].dropna().astype(str))
        check("REMARKS is a closed set matching the portal's options",
              vals <= {"PASSED", "FAILED"}, f"{sorted(vals)}")
        agree = (
            (pd.to_numeric(dfs["FINAL GRADE"], errors="coerce") >= 75)
            == (dfs["REMARKS"] == "PASSED")
        ).all()
        check("REMARKS agrees with the >= 75 rule on every row", bool(agree))

    print()
    if failures:
        print(f"{len(failures)} check(s) FAILED: {failures}")
        sys.exit(1)
    print("all checks passed")


if __name__ == "__main__":
    main()
